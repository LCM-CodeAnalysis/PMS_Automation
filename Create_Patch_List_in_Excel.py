import datetime as dt
from dateutil.relativedelta import relativedelta
import pandas as pd
from openpyxl import Workbook
import sys
import re
import requests
from bs4 import BeautifulSoup
import PMS_Data as pmsd
import time
import os

undecidedList = []

startPeriod = None
endPeriod = None

previousPatchList = []
previousPatchListPath = ''
importantSet = set()
criticalSet = set()
duplicationPatchList = []
newPatchList = []

def getPatchPeriod():
    global startPeriod
    global endPeriod
    today = dt.date.today()
    beforeOneMonth = today - relativedelta(months=1)
    startPeriod = getPatchDateByMonth(beforeOneMonth)
    endPeriod = getPatchDateByMonth(today)

def getPatchDateByMonth(dateTime):
    patchDate = dateTime.replace(day=1)
    weeks = 0
        
    while weeks < 2:
        if patchDate.weekday() == 1:
            weeks += 1
        patchDate += dt.timedelta(days=1)

    return patchDate

def isPatchExclusion(des):
    return any(one in des for one in pmsd.patchExclusionList)

def validatePatchInfo(guid, kbid, des):
    if not isinstance(kbid, float) or kbid == 0:
        return False
    if not isinstance(des, str) or isPatchExclusion(des):
        return False
    if (guid+'\t'+str(int(kbid))+'\n') in previousPatchList:
        duplicationPatchList.append([guid, kbid, des])
        return False
    return True

def makePreviousPatchList():
    global previousPatchList
    global startPeriod
    global previousPatchListPath

    previousPatchListPath = './' + startPeriod.strftime('%Y_%m_%d') + '_Previous_Patch_List.txt'
    f = open(previousPatchListPath, 'r')
    previousPatchList = f.readlines()
    f.close()


def makeSeveritySet():
    global importantSet
    global criticalSet
    global endPeriod

    xlsPath = './Security Updates ' + endPeriod.strftime('%Y-%m-%d') + '.csv'

    securityUpdates = pd.read_csv(xlsPath, encoding = 'ANSI')
    securityUpdates = securityUpdates.rename(columns={"Max Severity":"Severity"})

    for i in range(len(securityUpdates.Severity)):
        if not pd.isna(securityUpdates.Article[i]):
            if securityUpdates.Article[i].isdigit():
                if securityUpdates.Severity[i] is not None:
                    if('Important' == securityUpdates.Severity[i]):
                        importantSet.add(str(securityUpdates.Article[i]))
                    elif('Critical' == securityUpdates.Severity[i]):
                        criticalSet.add(str(securityUpdates.Article[i]))
            
    importantSet = importantSet.difference(criticalSet)

def setSeverity(excel, kbid):
    severityStr = ''
    if kbid in criticalSet:
        severityStr = '1'
    elif kbid in importantSet:
        severityStr = '0'
    return excel.replace('#s#', severityStr)

def getDownloadInfo(guid):
    url = 'https://catalog.update.microsoft.com/DownloadDialog.aspx'
    postData = {'updateIDs': '[{"size":0,"languages":"","uidInfo":"'+guid+'","updateID":"'+guid+'"}]'}
    regexForDownloadLink = "].url = '(https://.+)'"
    regexForFileName = "].fileName = '(.+)'"
    result = {}
    result['downloadLinkTuple'] = ()
    result['fileNameTuple'] = ()
    try:
        response = requests.request("POST", url, data=postData)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, "html.parser")
            headStr = str(soup.head)
            downloadLinkTuple = re.findall(regexForDownloadLink, headStr)
            result['downloadLinkTuple'] = downloadLinkTuple
            fileNameTuple = re.findall(regexForFileName, headStr)
            result['fileNameTuple'] = fileNameTuple
        else:
            None
    except requests.exceptions.Timeout as e:
        print("Timeout Error : ", e)
    except requests.exceptions.ConnectionError as e:
        print("Error Connecting : ", e)
    except requests.exceptions.HTTPError as e:
        print("Http Error : ", e)
    except requests.exceptions.RequestException as e:
        print("AnyException : ", e)

    return result

def addPatchRow(Classification, guid, kbid, des):
    global endPeriod
    regexList = pmsd.totalRegexDic[Classification]
    for regexDic in regexList:
        regexPattern = re.compile(regexDic['regex'].replace('#df0#', endPeriod.strftime('%Y-%m')))
        result = re.findall(regexPattern, des)
        length = len(result)
        if length == 0:
            continue
        elif length == 1:
            excel = regexDic['excel']
            descriptionInEnglish = regexDic['descriptionInEnglish']
            
            # KBID, GUID, dateForm1 적용
            excel = excel.replace('#ki#', kbid).replace('#gi#', guid).replace('#df1#', endPeriod.strftime('%Y-%m-%d'))
            
            # 개별 변경 사항 적용
            if isinstance(result[0], str):
                excel = excel.replace('#1#', result[0])
                descriptionInEnglish = descriptionInEnglish.replace('#1#', result[0])
            else:
                # type(result[0]) == tuple 일 경우
                for i in range(len(result[0])):
                    excel = excel.replace('#'+str(i+1)+'#', result[0][i])
                    descriptionInEnglish = descriptionInEnglish.replace('#'+str(i+1)+'#', result[0][i])
            
            # 심각도(Severity) 적용
            excel = setSeverity(excel, kbid)
            
            # multi_lan용 데이터 추출
            tempList = excel.split('\t')
            tempList[8] = endPeriod.strftime('%Y년 %m월, ') + tempList[8]
            name = tempList[1]
            descr_kor = tempList[8]
            refer = tempList[9]
            korList = [name, descr_kor, refer, '']
            descriptionInEnglish = endPeriod.strftime('%B, %Y ') + descriptionInEnglish
            enuList = [descriptionInEnglish, refer, '']

            # 다운로드 링크, 다운로드 파일 수 적용
            downloadInfo = getDownloadInfo(guid)
            downloadLinkTuple = downloadInfo['downloadLinkTuple']
            tempList.extend([des, len(downloadLinkTuple), ','.join(one for one in downloadLinkTuple)])

            # 최종행 저장
            if regexDic['group'] not in pmsd.totalRowDic[Classification]:
                pmsd.totalRowDic[Classification][regexDic['group']] = []
            tempList.extend(korList + enuList)
            pmsd.totalRowDic[Classification][regexDic['group']].append(tempList)
            return
        else:
            undecidedList.append([guid, kbid, des])
            return
    
    undecidedList.append([guid, kbid, des])

def addPatchRowForDotNet(guid, kbid, des):
    global endPeriod
    regexList = pmsd.totalRegexDic['dotnet']
    for regexDic in regexList:
        regexPattern = re.compile(regexDic['regex'])
        result = re.findall(regexPattern, des)
        length = len(result)
        if length == 0:
            continue
        elif length == 1:
            excel = regexDic['excel']
            descriptionInEnglish = regexDic['descriptionInEnglish']
            
            # KBID, GUID, dateForm1 적용
            excel = excel.replace('#ki#', kbid).replace('#gi#', guid).replace('#df1#', endPeriod.strftime('%Y-%m-%d'))
            
            # 개별 변경 사항 적용
            versionStr = ''
            if isinstance(result[0], str):
                excel = excel.replace('#1#', result[0])
                descriptionInEnglish = descriptionInEnglish.replace('#1#', result[0])
                if '.' in result[0]:
                    versionStr = result[0].replace('및', 'and').replace('  ', ', ')
                    excel = excel.replace('#dv#', versionStr)
            else:
                # type(result[0]) == tuple 일 경우
                for i in range(len(result[0])):
                    excel = excel.replace('#'+str(i+1)+'#', result[0][i])
                    descriptionInEnglish = descriptionInEnglish.replace('#'+str(i+1)+'#', result[0][i])
                    if '.' in result[0][i]:
                        versionStr = result[0][i].replace('및', 'and').replace('  ', ', ')
                        excel = excel.replace('#dv#', versionStr)
            
            # 심각도(Severity) 적용
            excel = setSeverity(excel, kbid)

            # multi_lan용 데이터 추출
            tempList = excel.split('\t')
            tempList[8] = endPeriod.strftime('%Y년 %m월, ') + tempList[8]
            name = tempList[1]
            descr_kor = tempList[8]
            refer = tempList[9]
            korList = [name, descr_kor, refer, '']
            descriptionInEnglish = endPeriod.strftime('%B, %Y ') + descriptionInEnglish.replace('#dv#', versionStr)
            enuList = [descriptionInEnglish, refer, '']

            # 다운로드 링크, 다운로드 파일 수 적용
            downloadInfo = getDownloadInfo(guid)
            downloadLinkTuple = downloadInfo['downloadLinkTuple']
            fileNameTuple = downloadInfo['fileNameTuple']

            downloadLength = len(downloadLinkTuple)
            # if downloadLength == 1:
            #     tempList.extend([des, 1, downloadLinkTuple[0]])
            #     # 최종행 저장
            #     if regexDic['group'] not in pmsd.totalRowDic['dotnet']:
            #         pmsd.totalRowDic['dotnet'][regexDic['group']] = []
            #     tempList.extend(korList + enuList)
            #     pmsd.totalRowDic['dotnet'][regexDic['group']].append(tempList)
            if downloadLength > 0:
                for i in range(len(downloadLinkTuple)):
                    copyList = tempList.copy()
                    copyList.extend([des, 1, downloadLinkTuple[i]])
                    for fileNameRegexDic in regexDic['fileName']:
                        fileNameRegexPattern = re.compile(fileNameRegexDic['regex'])
                        resultForFileName = re.findall(fileNameRegexPattern, fileNameTuple[i])
                        length = len(resultForFileName)
                        if length == 0:
                            continue
                        elif length == 1:
                            excelForFileName = fileNameRegexDic['excel']
                            if isinstance(result[0], str):
                                excelForFileName = excelForFileName.replace('#1#', result[0])
                            else:
                                for i in range(len(result[0])):
                                    excelForFileName = excelForFileName.replace('#'+str(i+1)+'#', result[0][i])
                            if isinstance(resultForFileName[0], str):
                                excelForFileName = excelForFileName.replace('#f1#', resultForFileName[0])
                            else:
                                for i in range(len(resultForFileName[0])):
                                    excelForFileName = excelForFileName.replace('#f'+str(i+1)+'#', resultForFileName[0][i])
                            copyList[16] = excelForFileName
                    #최종행 저장
                    if regexDic['group'] not in pmsd.totalRowDic['dotnet']:
                        pmsd.totalRowDic['dotnet'][regexDic['group']] = []
                    copyList.extend(korList + enuList)
                    pmsd.totalRowDic['dotnet'][regexDic['group']].append(copyList)
            else:
                tempList.extend([des, 0, ''])
                # 최종행 저장
                if regexDic['group'] not in pmsd.totalRowDic['dotnet']:
                    pmsd.totalRowDic['dotnet'][regexDic['group']] = []
                tempList.extend(korList + enuList)
                pmsd.totalRowDic['dotnet'][regexDic['group']].append(tempList)
            return
        else:
            undecidedList.append([guid, kbid, des])
            return
    
    undecidedList.append([guid, kbid, des])

def createPatchRowsByType(guid, kbid, des):
    if '.Net' in des or '.NET' in des:
        addPatchRowForDotNet(guid, kbid, des)
    elif 'Azure' in des:
        addPatchRow('azure', guid, kbid, des)
    elif 'Internet' in des:
        addPatchRow('internet', guid, kbid, des)
    elif 'Windows' in des:
        if any(one in des for one in ['누적', 'Cumulative']):
            addPatchRow('windows-cumulative', guid, kbid, des)
        elif any(one in des for one in ['보안', 'Security']):
            addPatchRow('windows-security', guid, kbid, des)
        else:
            addPatchRow('windows-etc', guid, kbid, des)
    elif 'Exchange' in des:
        addPatchRow('exchange', guid, kbid, des)
    elif 'PowerShell' in des:
        addPatchRow('powershell', guid, kbid, des)
    elif any(one in des for one in pmsd.officeList):
        addPatchRow('office', guid, kbid, des)
    else:
        addPatchRow('etc', guid, kbid, des)

def readPatchListFromExcel():
    global endPeriod
    xlsPath = './' + endPeriod.strftime('%Y_%m_%d') + '_Result.csv'
    patchList = pd.read_csv(xlsPath, encoding = 'ANSI', names=['day', 'GUID', 'c', 'd', 'KBID', 'Des'])
    if patchList.shape[0] < 1: # == len(patchTargetList)
        print('패치 목록을 불러오는데 실패했습니다.')
        sys.exit()
    else:
        return patchList
    
def createPatchRows(patchList):
    global startPeriod
    global endPeriod
    for i in reversed(range(patchList.shape[0])):
        try:
            row_datetime = dt.datetime.strptime(patchList.day[i], '%Y-%m-%dT%H:%M:%SZ').date()
            if row_datetime >= endPeriod:
                continue
            elif row_datetime < startPeriod:
                break
            else:
                if validatePatchInfo(patchList.GUID[i], patchList.KBID[i], patchList.Des[i]):
                    createPatchRowsByType(patchList.GUID[i], str(int(patchList.KBID[i])), patchList.Des[i])
                    newPatchList.append(patchList.GUID[i]+'\t'+str(int(patchList.KBID[i]))+'\n')
        except ValueError as e: # 날짜영역에 문자열이 들어있는 경우
            print('ValueError : ' ,e)
        except TypeError as e:  # 날짜영역이 비어있는 경우
            print('TypeError : ', e)

    return None

def writePatchListToExcel():
    global endPeriod
    wb = Workbook()
    normal_ws = wb.active
    normal_ws.title = 'normal'
    for dic in pmsd.totalRowDic.values():
        dic = dict(sorted(dic.items()))
        for list in dic.values():
            list.sort(key=lambda x:x[8])
            for one in list:
                normal_ws.append(one)
            normal_ws.append([])
            normal_ws.append([])

    exception_ws = wb.create_sheet()
    exception_ws.title = 'exception'
    for one in undecidedList:
        exception_ws.append(one)

    duplication_ws =wb.create_sheet()
    duplication_ws.title = 'duplication'
    for one in duplicationPatchList:
        duplication_ws.append(one)

    xlsxPath = './' + endPeriod.strftime('%Y_%m_%d') + '_Auto_Patch.xlsx'
    wb.save(xlsxPath)

def updatePreviousPatchList():
    f = open(previousPatchListPath, 'a')
    for one in newPatchList:
        f.write(one)
    f.close()
    os.rename(previousPatchListPath, './' + endPeriod.strftime('%Y_%m_%d') + '_Previous_Patch_List.txt')

def main():
    global startPeriod
    global endPeriod
    numberOfArgs = len(sys.argv)
    if numberOfArgs == 1:
        getPatchPeriod()
    elif numberOfArgs == 3:
        startPeriod, endPeriod = dt.datetime.strptime(sys.argv[1], '%Y%m%d').date(), dt.datetime.strptime(sys.argv[2], '%Y%m%d').date()
    else:
        print('파라미터 개수를 확인해주세요.')
        sys.exit()

    makeSeveritySet()
    makePreviousPatchList()

    patchList = readPatchListFromExcel()
    createPatchRows(patchList)
    writePatchListToExcel()

    updatePreviousPatchList()

startTime = time.time()
main()
endTime = time.time()
print(f'실행시간 : {endTime-startTime:.3f} sec')